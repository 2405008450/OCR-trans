import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from memory import memory as memory_module


def test_coverage_accepts_whitespace_only_difference():
    rows = [{"原文": "Hello world", "译文": "你好 世界"}]

    accepted, info = memory_module.validate_alignment_coverage_only(
        rows,
        "Hello\nworld",
        "你好   世界",
        report_path=None,
    )

    assert accepted == [{"原文": "Hello\nworld", "译文": "你好   世界"}]
    assert info["rejected"] == 0


def test_coverage_rejects_modified_text():
    rows = [{"原文": "Hello brave world", "译文": "你好世界"}]

    accepted, info = memory_module.validate_alignment_coverage_only(
        rows,
        "Hello world",
        "你好世界",
        report_path=None,
    )

    assert accepted == []
    assert info["rejected"] == 1
    assert "Hello world" in info["remaining_original"]


def test_coverage_consumes_duplicate_text_in_order():
    rows = [
        {"原文": "A", "译文": "甲"},
        {"原文": "A", "译文": "甲"},
    ]

    accepted, info = memory_module.validate_alignment_coverage_only(
        rows,
        "A\nA",
        "甲\n甲",
        report_path=None,
    )

    assert accepted == rows
    assert info["rejected"] == 0
    assert info["remaining_original"] == ""
    assert info["remaining_translated"] == ""


def test_coverage_ignores_invisible_format_chars():
    rows = [{"原文": "财金〔2023〕100号", "译文": "C.J. [2023] No. 100"}]

    accepted, info = memory_module.validate_alignment_coverage_only(
        rows,
        "财金〔\u200b2023〕\ufeff100号",
        "C.J. [2023]\u00ad No. 100",
        report_path=None,
    )

    assert accepted == [{"原文": "财金〔\u200b2023〕\ufeff100号", "译文": "C.J. [2023]\u00ad No. 100"}]
    assert info["rejected"] == 0


def test_coverage_marks_already_consumed_text():
    original_consumer = memory_module.CoverageTextConsumer("A")
    translated_consumer = memory_module.CoverageTextConsumer("甲甲")

    accepted, rejected = memory_module._consume_alignment_rows(
        [
            {"原文": "A", "译文": "甲"},
            {"原文": "A", "译文": "甲"},
        ],
        original_consumer,
        translated_consumer,
        "测试阶段",
        "测试来源",
    )

    assert accepted == [{"原文": "A", "译文": "甲"}]
    assert rejected[0]["问题类型"] == "原文已被前序行消费"
    assert rejected[0]["原文命中"] == "已被前序行消费"


def test_final_repair_appends_only_verified_rows(monkeypatch):
    def fake_call_llm_stream(system_prompt, user_prompt, model_id, filename=""):
        return "B ||| 乙\n幻觉 ||| hallucination"

    monkeypatch.setattr(memory_module, "call_llm_stream", fake_call_llm_stream)

    final_rows, report = memory_module._repair_alignment_rows_with_full_text(
        [{"原文": "A", "译文": "甲"}],
        "A\nB",
        "甲\n乙",
        "fake-model",
        report_path=None,
    )

    assert final_rows == [
        {"原文": "A", "译文": "甲"},
        {"原文": "B", "译文": "乙"},
    ]
    assert any(row["原文"] == "幻觉" for row in report["rejected"])
    assert report["remaining_original"] == ""
    assert report["remaining_translated"] == ""


def test_final_repair_appended_rows_can_be_highlighted(monkeypatch, tmp_path):
    def fake_call_llm_stream(system_prompt, user_prompt, model_id, filename=""):
        return "B ||| 乙\nC ||| 丙"

    monkeypatch.setattr(memory_module, "call_llm_stream", fake_call_llm_stream)

    final_rows, report = memory_module._repair_alignment_rows_with_full_text(
        [{"原文": "A", "译文": "甲"}],
        "A\nB\nC",
        "甲\n乙\n丙",
        "fake-model",
        report_path=None,
    )

    assert report["appended_row_indices"] == [1, 2]

    excel_path = tmp_path / "alignment.xlsx"
    pd.DataFrame(final_rows, columns=["原文", "译文"]).to_excel(excel_path, index=False)
    memory_module._highlight_final_repair_rows(excel_path, report["appended_row_indices"])

    ws = load_workbook(excel_path).active
    repair_fill = memory_module.FINAL_REPAIR_FILL_COLOR
    assert ws.cell(row=2, column=1).fill.start_color.rgb != f"00{repair_fill}"
    assert ws.cell(row=3, column=1).fill.start_color.rgb == f"00{repair_fill}"
    assert ws.cell(row=4, column=2).fill.start_color.rgb == f"00{repair_fill}"
