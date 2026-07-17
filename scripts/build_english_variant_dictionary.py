# -*- coding: utf-8 -*-
"""把英美式英语 Excel 词库编译为运行时 JSON。"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = REPO_ROOT / "data" / "english_variant" / "英美式英语词汇对比_名词_260505.xlsx"
DEFAULT_OUTPUT = REPO_ROOT / "data" / "english_variant" / "dictionary.json"

NOUN_HEADERS = ("英式单数", "英式复数", "美式单数", "美式复数")
VERB_HEADERS = (
    "英式原型",
    "英式第三人称单数",
    "英式过去时",
    "英式过去分词",
    "英式ing",
    "美式原型",
    "美式第三人称单数",
    "美式过去时",
    "美式过去分词",
    "美式ing",
)


def _clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def _source_ref(sheet: str, row: int, form: str) -> dict[str, Any]:
    return {"sheet": sheet, "row": row, "form": form}


def _iter_pairs(source: Path) -> Iterable[dict[str, Any]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        required_sheets = {"名词", "动词"}
        missing = required_sheets.difference(workbook.sheetnames)
        if missing:
            raise ValueError(f"词库缺少工作表: {', '.join(sorted(missing))}")

        noun_sheet = workbook["名词"]
        noun_headers = tuple(_clean(cell.value) for cell in noun_sheet[1][:4])
        if noun_headers != NOUN_HEADERS:
            raise ValueError(f"名词工作表表头不正确: {noun_headers!r}")
        noun_forms = ((0, 2, "singular"), (1, 3, "plural"))
        for row_index, row in enumerate(noun_sheet.iter_rows(min_row=2, values_only=True), start=2):
            for british_col, american_col, form in noun_forms:
                british = _clean(row[british_col] if british_col < len(row) else None)
                american = _clean(row[american_col] if american_col < len(row) else None)
                if british and american and british.casefold() != american.casefold():
                    yield {
                        "british": british,
                        "american": american,
                        "source": _source_ref("名词", row_index, form),
                    }

        verb_sheet = workbook["动词"]
        verb_headers = tuple(_clean(cell.value) for cell in verb_sheet[1][:10])
        if verb_headers != VERB_HEADERS:
            raise ValueError(f"动词工作表表头不正确: {verb_headers!r}")
        verb_forms = ("base", "third_person", "past", "past_participle", "ing")
        for row_index, row in enumerate(verb_sheet.iter_rows(min_row=2, values_only=True), start=2):
            for index, form in enumerate(verb_forms):
                british = _clean(row[index] if index < len(row) else None)
                american_index = index + 5
                american = _clean(row[american_index] if american_index < len(row) else None)
                if british and american and british.casefold() != american.casefold():
                    yield {
                        "british": british,
                        "american": american,
                        "source": _source_ref("动词", row_index, form),
                    }
    finally:
        workbook.close()


def _compile_direction(
    pairs: list[dict[str, Any]],
    source_field: str,
    target_field: str,
) -> dict[str, Any]:
    grouped: dict[str, dict[str, Any]] = {}
    for pair in pairs:
        source_text = pair[source_field]
        target_text = pair[target_field]
        source_key = source_text.casefold()
        target_key = target_text.casefold()
        source_entry = grouped.setdefault(
            source_key,
            {"source": source_text, "targets": {}},
        )
        target_entry = source_entry["targets"].setdefault(
            target_key,
            {"target": target_text, "sources": []},
        )
        if pair["source"] not in target_entry["sources"]:
            target_entry["sources"].append(pair["source"])

    rules: list[dict[str, Any]] = []
    ambiguous: list[dict[str, Any]] = []
    for source_key in sorted(grouped):
        entry = grouped[source_key]
        targets = [entry["targets"][key] for key in sorted(entry["targets"])]
        for target in targets:
            target["sources"].sort(key=lambda item: (item["sheet"], item["row"], item["form"]))
        if len(targets) == 1:
            rules.append(
                {
                    "source": entry["source"],
                    "target": targets[0]["target"],
                    "sources": targets[0]["sources"],
                }
            )
        else:
            ambiguous.append(
                {
                    "source": entry["source"],
                    "candidates": targets,
                }
            )
    return {"rules": rules, "ambiguous": ambiguous}


def compile_dictionary(source: Path = DEFAULT_SOURCE) -> dict[str, Any]:
    source = source.resolve()
    if not source.is_file():
        raise FileNotFoundError(f"词库文件不存在: {source}")

    pairs = list(_iter_pairs(source))
    if not pairs:
        raise ValueError("词库中没有可用的英美词对")

    content = source.read_bytes()
    source_sha256 = hashlib.sha256(content).hexdigest()
    version_match = re.search(r"(\d{6,8})", source.stem)
    dictionary_version = version_match.group(1) if version_match else source_sha256[:12]
    generated_at = datetime.fromtimestamp(source.stat().st_mtime, tz=timezone.utc).isoformat()

    british_to_american = _compile_direction(pairs, "british", "american")
    american_to_british = _compile_direction(pairs, "american", "british")
    unique_pairs = {
        (pair["british"].casefold(), pair["american"].casefold()) for pair in pairs
    }

    return {
        "schema_version": 1,
        "dictionary_version": dictionary_version,
        "source_file": source.name,
        "source_sha256": source_sha256,
        "generated_at": generated_at,
        "stats": {
            "raw_pairs": len(pairs),
            "unique_pairs": len(unique_pairs),
            "british_to_american_rules": len(british_to_american["rules"]),
            "british_to_american_ambiguous": len(british_to_american["ambiguous"]),
            "american_to_british_rules": len(american_to_british["rules"]),
            "american_to_british_ambiguous": len(american_to_british["ambiguous"]),
        },
        "directions": {
            "british_to_american": british_to_american,
            "american_to_british": american_to_british,
        },
    }


def render_dictionary(payload: dict[str, Any]) -> str:
    return json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=False) + "\n"


def build_dictionary(
    source: Path = DEFAULT_SOURCE,
    output: Path = DEFAULT_OUTPUT,
    *,
    check: bool = False,
) -> bool:
    rendered = render_dictionary(compile_dictionary(source))
    if check:
        return output.is_file() and output.read_text(encoding="utf-8") == rendered
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(rendered, encoding="utf-8")
    return True


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--check", action="store_true", help="只检查 JSON 是否为最新")
    args = parser.parse_args()

    ok = build_dictionary(args.source, args.output, check=args.check)
    if args.check and not ok:
        print(f"运行时词库不是最新版本: {args.output}")
        return 1
    payload = compile_dictionary(args.source)
    print(json.dumps(payload["stats"], ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
