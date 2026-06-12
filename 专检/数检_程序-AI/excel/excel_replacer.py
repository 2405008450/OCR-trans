"""
Excel 替换与批注模块
使用 openpyxl 在工作表单元格中查找文本、替换并添加批注。
支持：上下文验证、文本变体容错、Unicode/不可见字符清洗。
"""
import openpyxl
from openpyxl.comments import Comment
from pathlib import Path
from typing import Optional, List, Tuple, Dict
from pdf.text_matcher import TextMatcher, clean_text_thoroughly, generate_search_variants


class ExcelReplacer:
    """Excel 替换器 — 遍历所有工作表的单元格，执行文本替换并添加批注"""

    def __init__(self, xlsx_path: str):
        self.xlsx_path = Path(xlsx_path)
        if not self.xlsx_path.exists():
            raise FileNotFoundError(f"文件不存在: {xlsx_path}")
        self.wb = openpyxl.load_workbook(str(self.xlsx_path))
        self.replacements_count = 0
        self.annotations_count = 0
        self._matcher = TextMatcher()

    # ── 公开接口 ──

    def replace_and_annotate(
        self,
        old_text: str,
        new_text: str,
        reason: str = "",
        context: str = "",
        highlight: bool = False,
        prev_tgt: str = "",
        next_tgt: str = "",
    ) -> bool:
        """
        在所有工作表中查找包含 old_text 的单元格，替换为 new_text。
        多候选时优先用前后句译文夹逼定位，其次用上下文排序。
        """
        candidates = self._find_all_candidates(old_text)
        if not candidates:
            return False

        # 多候选：先用前后句夹逼，再用上下文排序
        if len(candidates) > 1:
            if prev_tgt or next_tgt:
                sandwiched = self._sandwich_by_neighbors(candidates, prev_tgt, next_tgt)
                if sandwiched:
                    candidates = sandwiched
                elif context:
                    candidates = self._rank_by_context(candidates, context)
            elif context:
                candidates = self._rank_by_context(candidates, context)

        # 执行替换（最佳匹配）
        best = candidates[0]
        cell = best["cell"]
        cell_str = str(cell.value)
        matched = best["matched_text"]
        strategy = best["strategy"]

        # 执行替换
        if matched in cell_str:
            cell.value = cell_str.replace(matched, new_text, 1)
        else:
            # 清洗后替换
            ok = self._replace_fuzzy_in_cell(cell, old_text, new_text)
            if not ok:
                return False

        self.replacements_count += 1

        # 精确高亮：只对新文本部分设红色字体（富文本）
        if highlight:
            self._highlight_new_text_in_cell(cell, new_text)

        # 添加批注
        comment_text = f"[修改] '{old_text}' → '{new_text}'"
        if reason:
            comment_text += f"\n理由: {reason}"
        comment_text += f"\n匹配策略: {strategy}"

        if cell.comment:
            cell.comment.text += "\n" + comment_text
        else:
            cell.comment = Comment(comment_text, "标点检查")
        self.annotations_count += 1

        return True

    @staticmethod
    def _highlight_new_text_in_cell(cell, new_text: str):
        """
        用富文本对单元格内的 new_text 部分设红色字体，其余保持原样。
        Excel 不支持部分背景色，用字体颜色标注。
        """
        try:
            from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont
            cell_str = str(cell.value) if cell.value is not None else ""
            if not new_text or new_text not in cell_str:
                return
            idx = cell_str.find(new_text)
            prefix = cell_str[:idx]
            suffix = cell_str[idx + len(new_text):]

            red_font = InlineFont(color="FF0000", b=True)
            parts = []
            if prefix:
                parts.append(prefix)
            parts.append(TextBlock(red_font, new_text))
            if suffix:
                parts.append(suffix)

            cell.value = CellRichText(*parts)
        except Exception:
            pass

    def save(self, output_path: Optional[str] = None) -> str:
        out = Path(output_path) if output_path else self.xlsx_path
        self.wb.save(str(out))
        print(f"✓ 已保存 Excel: {out}")
        print(f"  替换数量: {self.replacements_count}")
        print(f"  批注数量: {self.annotations_count}")
        return str(out)

    # ── 智能查找逻辑 ──

    def _find_all_candidates(self, old_text: str) -> List[Dict]:
        """
        在所有工作表中查找 old_text 的所有候选位置。
        多层匹配：精确 → 清洗后 → 变体 → 模糊。
        """
        candidates = []
        variants = generate_search_variants(old_text)
        old_clean = clean_text_thoroughly(old_text)

        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    cell_str = str(cell.value)
                    if not cell_str.strip():
                        continue

                    match_result = self._try_match(cell_str, old_text, old_clean, variants)
                    if match_result:
                        candidates.append({
                            "cell": cell,
                            "ws_title": ws.title,
                            "cell_ref": cell.coordinate,
                            "cell_text": cell_str,
                            "matched_text": match_result[0],
                            "strategy": match_result[1],
                            "score": match_result[2],
                        })

        candidates.sort(key=lambda c: c["score"], reverse=True)
        return candidates

    def _try_match(
        self, cell_str: str, old_text: str, old_clean: str, variants: List[str]
    ) -> Optional[Tuple[str, str, float]]:
        """
        多层匹配策略，返回 (实际匹配文本, 策略名, 分数) 或 None。
        """
        # 1. 精确匹配
        if old_text in cell_str:
            return (old_text, "精确匹配", 1.0)

        cell_clean = clean_text_thoroughly(cell_str)

        # 2. 清洗后匹配
        if old_clean and old_clean in cell_clean:
            return (old_text, "清洗后匹配", 0.9)

        # 3. 忽略空格匹配
        cell_no_sp = cell_str.replace(" ", "").replace("\u3000", "").replace("\xa0", "")
        old_no_sp = old_text.replace(" ", "").replace("\u3000", "").replace("\xa0", "")
        if old_no_sp and old_no_sp in cell_no_sp:
            return (old_text, "忽略空格匹配", 0.85)

        # 4. 变体匹配
        for v in variants:
            if v == old_text:
                continue
            if v in cell_str:
                return (v, f"变体匹配({v})", 0.7)
            v_clean = clean_text_thoroughly(v)
            if v_clean and v_clean in cell_clean:
                return (v, f"变体清洗匹配({v})", 0.65)

        # 5. TextMatcher 7层匹配
        found, start, end, strategy = self._matcher.find_best_match(
            cell_str, old_text, return_strategy=True
        )
        if found and strategy and "精确" not in strategy:
            return (old_text, f"TextMatcher:{strategy}", 0.5)

        return None

    def _sandwich_by_neighbors(self, candidates: List[Dict], prev_tgt: str, next_tgt: str) -> List[Dict]:
        """
        用前一句/后一句译文在工作表中精确定位，夹逼出正确的候选单元格。
        策略：
          1. 全表精确搜索前句/后句所在单元格的行号
          2. 找满足 prev_row < candidate_row < next_row 的候选
          3. 唯一时直接返回；多个时取行距最近的
        """
        from pdf.text_matcher import clean_text_thoroughly as _clean

        prev_clean = _clean(prev_tgt) if prev_tgt else ""
        next_clean = _clean(next_tgt) if next_tgt else ""

        if not prev_clean and not next_clean:
            return []

        # 构建全表文本索引：(ws_title, row) -> clean_text
        row_texts: dict = {}
        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                row_parts = []
                for cell in row:
                    if cell.value:
                        row_parts.append(str(cell.value))
                if row_parts:
                    key = (ws.title, row[0].row)
                    row_texts[key] = _clean(" ".join(row_parts))

        def _find_row_positions(target_clean: str) -> list:
            if not target_clean:
                return []
            exact = [k for k, v in row_texts.items() if target_clean in v]
            if exact:
                return exact
            # 降级：字符覆盖率 >= 0.8
            result = []
            for k, v in row_texts.items():
                if not v:
                    continue
                overlap = sum(1 for ch in target_clean if ch in v) / len(target_clean)
                if overlap >= 0.8:
                    result.append(k)
            return result

        prev_positions = _find_row_positions(prev_clean)
        next_positions = _find_row_positions(next_clean)

        # 候选的 (ws_title, row)
        def _candidate_key(c: dict):
            return (c["ws_title"], c["cell"].row)

        MAX_GAP = 30

        def _is_sandwiched(c: dict) -> bool:
            ws, row = _candidate_key(c)
            # 同一 sheet 的前后句才有意义
            prev_same = [(w, r) for w, r in prev_positions if w == ws]
            next_same = [(w, r) for w, r in next_positions if w == ws]
            if prev_same and not next_same:
                return any(r < row and row - r <= MAX_GAP for _, r in prev_same)
            if next_same and not prev_same:
                return any(row < r and r - row <= MAX_GAP for _, r in next_same)
            closest_prev = max((r for w, r in prev_same if r < row), default=None)
            closest_next = min((r for w, r in next_same if r > row), default=None)
            if closest_prev is None or closest_next is None:
                return False
            return (row - closest_prev) <= MAX_GAP and (closest_next - row) <= MAX_GAP

        sandwiched = [c for c in candidates if _is_sandwiched(c)]
        if not sandwiched:
            return []

        if len(sandwiched) == 1:
            return sandwiched

        # 多个：取行距最近的
        def _gap(c: dict) -> int:
            ws, row = _candidate_key(c)
            prev_same = [(w, r) for w, r in prev_positions if w == ws]
            next_same = [(w, r) for w, r in next_positions if w == ws]
            cp = max((r for w, r in prev_same if r < row), default=row)
            cn = min((r for w, r in next_same if r > row), default=row)
            return (row - cp) + (cn - row)

        sandwiched.sort(key=_gap)
        best_gap = _gap(sandwiched[0])
        # 最近和次近距离相同 → 无法区分，返回空交由上下文排序
        if len(sandwiched) > 1 and _gap(sandwiched[1]) == best_gap:
            return []
        return [sandwiched[0]]

    def _rank_by_context(self, candidates: List[Dict], context: str) -> List[Dict]:
        """根据上下文对候选项排序"""
        context_clean = clean_text_thoroughly(context)

        for c in candidates:
            # Excel 上下文：同一行的所有单元格文本
            surrounding = self._get_surrounding_text(c)
            surrounding_clean = clean_text_thoroughly(surrounding)

            if context_clean and surrounding_clean:
                overlap = self._char_overlap(context_clean, surrounding_clean)
            else:
                overlap = 0.0

            c["context_score"] = c["score"] * 0.4 + overlap * 0.6

        candidates.sort(key=lambda c: c["context_score"], reverse=True)
        return candidates

    def _get_surrounding_text(self, candidate: Dict) -> str:
        """获取候选单元格所在行的所有文本"""
        cell = candidate["cell"]
        ws = cell.parent
        row_num = cell.row
        parts = []
        for row_cell in ws[row_num]:
            if row_cell.value is not None:
                parts.append(str(row_cell.value))
        return " ".join(parts)

    @staticmethod
    def _char_overlap(text1: str, text2: str) -> float:
        """字符重叠率"""
        if not text1 or not text2:
            return 0.0
        shorter, longer = (text1, text2) if len(text1) <= len(text2) else (text2, text1)
        match_count = sum(1 for ch in shorter if ch in longer)
        return match_count / len(shorter) if shorter else 0.0

    # ── 替换辅助 ──

    def _replace_fuzzy_in_cell(self, cell, old_text: str, new_text: str) -> bool:
        """当精确替换失败时，用清洗/模糊方式定位并替换"""
        cell_str = str(cell.value)

        # TextMatcher 定位
        found, start, end, _ = self._matcher.find_best_match(cell_str, old_text)
        if found and start is not None and end is not None:
            cell.value = cell_str[:start] + new_text + cell_str[end:]
            return True

        # 忽略空格定位
        cell_no_sp = cell_str.replace(" ", "")
        old_no_sp = old_text.replace(" ", "")
        if old_no_sp in cell_no_sp:
            idx = cell_no_sp.index(old_no_sp)
            orig_start = self._map_no_space_pos(cell_str, idx)
            orig_end = self._map_no_space_pos(cell_str, idx + len(old_no_sp))
            cell.value = cell_str[:orig_start] + new_text + cell_str[orig_end:]
            return True

        return False

    @staticmethod
    def _map_no_space_pos(text: str, no_space_pos: int) -> int:
        """将去空格后的位置映射回原文位置"""
        count = 0
        for i, ch in enumerate(text):
            if ch != " ":
                if count == no_space_pos:
                    return i
                count += 1
        return len(text)
