"""
中英对照对齐模块

run() 返回 List[Dict]（JSON 可序列化），不含任何 IO。
调用方拿到数据后自行决定输出方式（Excel / JSON 文件 / 传给下游）。
"""
import os
import json
import hashlib
from typing import List, Dict
from openai import OpenAI
from dotenv import load_dotenv

from body_extractor import extract_body_text
from text_splitter import split_text_pair, get_overlap_text

load_dotenv()
client = OpenAI(
    api_key=os.getenv("API_KEY"),
    base_url=os.getenv("BASE_URL"),
)


# ─── 工具 ────────────────────────────────────────────────────

def _make_hash(src: str, tgt: str) -> str:
    return hashlib.md5((src + "||" + tgt).encode("utf-8")).hexdigest()


# ─── LLM 对齐单块 ────────────────────────────────────────────

def _align_chunk(chunk_id: int, src: str, tgt: str, seen: set,
                 src_overlap: str = "", tgt_overlap: str = "") -> List[Dict]:
    """对单个 chunk 调用 LLM 对齐，返回去重后的对齐结果列表。

    每项：{"chunk_id": int, "source": str, "target": str, "hash": str}
    """
    overlap_hint = ""
    if src_overlap or tgt_overlap:
        overlap_hint = f"""
⚠️ 以下内容是与上一块的重叠段落，已经对齐过，请直接跳过，不要再次输出：
【已对齐原文（跳过）】:
{src_overlap}

【已对齐译文（跳过）】:
{tgt_overlap}

"""

    prompt = f"""你是专业中英对齐系统。

⚠️重要规则：
1. 只对当前chunk内容进行对齐
2. 不允许使用 overlap 内容生成新的对齐
3. 不得重复已出现内容
4. 输出必须是 JSON 数组
{overlap_hint}
输出格式：
[{{"source": "...", "target": "..."}}]

原文：
{src}

译文：
{tgt}
"""
    try:
        resp = client.chat.completions.create(
            model="google/gemini-3-flash-preview",
            messages=[
                {"role": "system", "content": "严格对齐系统"},
                {"role": "user",   "content": prompt},
            ],
            temperature=0,
        )
        content = resp.choices[0].message.content
        content = content.replace("```json", "").replace("```", "")
        data = json.loads(content)

        results = []
        for item in data:
            s = item.get("source", "").strip()
            t = item.get("target", "").strip()
            if not s and not t:
                continue
            h = _make_hash(s, t)
            if h in seen:
                continue
            seen.add(h)
            results.append({"chunk_id": chunk_id, "source": s, "target": t, "hash": h})

        return results

    except Exception as e:
        print(f"❌ chunk {chunk_id} 失败: {e}")
        return []


# ─── 主流程（返回 JSON） ──────────────────────────────────────

def run(src_path: str, tgt_path: str) -> List[Dict]:
    """
    读取原文/译文 docx → 分块 → LLM 对齐 → 返回 List[Dict]。

    每项结构：
    {
        "chunk_id": int,
        "source":   str,   # 原文句
        "target":   str,   # 译文句
        "hash":     str,
    }
    """
    print("📄 读取文档...")
    src_text = extract_body_text(src_path)
    tgt_text = extract_body_text(tgt_path)

    print("✂️ 分块（带overlap）...")
    chunks = split_text_pair(src_text, tgt_text)
    print(f"📦 chunk数量: {len(chunks)}")

    seen_hash: set = set()
    all_results: List[Dict] = []

    print("🤖 LLM逐块对齐...")
    for i, (src_chunk, tgt_chunk, *_) in enumerate(chunks):
        print(f"  ➡️ chunk {i + 1}/{len(chunks)}")
        # 计算与上一块的实际重叠文本，告知 LLM 跳过
        src_overlap, tgt_overlap = "", ""
        if i > 0:
            src_overlap, tgt_overlap = get_overlap_text(chunks[i - 1], chunks[i])
        results = _align_chunk(i, src_chunk, tgt_chunk, seen_hash, src_overlap, tgt_overlap)
        all_results.extend(results)

    print(f"� 总对齐结果: {len(all_results)}")
    return all_results


# ─── 入口 ────────────────────────────────────────────────────

if __name__ == "__main__":
    from report_generator import generate_align_report

    src_path = r"C:\Users\H\Desktop\数检_程序-AI\测试文件\原文-含不可编辑_01 (2026-007)2025年年度报告.docx"
    tgt_path = r"C:\Users\H\Desktop\数检_程序-AI\测试文件\译文-含不可编辑_01 (2026-007)2025年年度报告(1).docx"

    rows = run(src_path, tgt_path)
    generate_align_report(rows, output_path="reports/final_align.xlsx")
"""
报告生成模块

接收 List[Dict]（JSON 数据），内部转 DataFrame 后写入 Excel。
"""
import os
from datetime import datetime
from typing import List, Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────
# 内部工具
# ─────────────────────────────────────────

def _ensure_dir(path: str):
    d = os.path.dirname(path)
    if d:
        os.makedirs(d, exist_ok=True)


def _style_header(ws, fill_hex: str):
    fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _set_col_widths(ws, widths: List[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_df(ws, df: pd.DataFrame, fill_hex: str, col_widths: List[int] = None):
    ws.append(list(df.columns))
    _style_header(ws, fill_hex)
    for row in df.itertuples(index=False):
        ws.append(list(row))
    if col_widths:
        _set_col_widths(ws, col_widths)


def _to_df(rows: List[Dict], cols: List[str]) -> pd.DataFrame:
    """从 JSON rows 中取指定列，列表字段转逗号字符串。"""
    df = pd.DataFrame(rows)
    # 只保留存在的列
    df = df[[c for c in cols if c in df.columns]].copy()
    for c in ["原文数值", "译文数值"]:
        if c in df.columns:
            df[c] = df[c].apply(lambda x: ", ".join(x) if isinstance(x, list) else x)
    return df


# ─────────────────────────────────────────
# 合并报告列定义
# ─────────────────────────────────────────

_COLS = [
    "原文", "译文", "原文数值", "译文数值",
    "是否错误", "错误类型", "错误原因",
    "AI是否正确", "AI错误数量", "AI错误类型", "AI错误详情",
    "一致性", "差异类型",
]
_WIDTHS = [60, 60, 22, 22, 10, 18, 55, 12, 12, 18, 50, 10, 12]


# ─────────────────────────────────────────
# 1. 程序+AI 合并报告（主入口）
# ─────────────────────────────────────────

def generate_combined_report(rows: List[Dict],
                              output_path: str = "reports/final_checked.xlsx"):
    """
    接收 merge_ai_results() 返回的 List[Dict]，生成合并 Excel。

    Sheet 说明：
      全量结果  — 所有行，规则列 + AI列并排
      确认错误  — 规则 & AI 均报错
      漏检FN    — 规则正确但 AI 报错
      误报FP    — 规则报错但 AI 认为正确
      统计汇总  — 各类数量
    """
    _ensure_dir(output_path)
    wb = Workbook()
    wb.remove(wb.active)

    df = pd.DataFrame(rows)

    has_ai   = "AI是否正确" in df.columns
    has_diff = "差异类型"   in df.columns

    # Sheet1: 全量
    _write_df(wb.create_sheet("全量结果"), _to_df(rows, _COLS), "2E75B6", _WIDTHS)

    # Sheet2: 确认错误
    confirmed_rows = [r for r in rows
                      if r.get("是否错误") == "❗错误" and r.get("AI是否正确") == "❗错误"]
    if confirmed_rows:
        _write_df(wb.create_sheet("确认错误"), _to_df(confirmed_rows, _COLS), "C00000", _WIDTHS)

    # Sheet3: 漏检 FN
    fn_rows = [r for r in rows if r.get("差异类型") == "漏检（FN）"]
    if fn_rows:
        _write_df(wb.create_sheet("漏检FN"), _to_df(fn_rows, _COLS), "ED7D31", _WIDTHS)

    # Sheet4: 误报 FP
    fp_rows = [r for r in rows if r.get("差异类型") == "误报（FP）"]
    if fp_rows:
        _write_df(wb.create_sheet("误报FP"), _to_df(fp_rows, _COLS), "7030A0", _WIDTHS)

    # Sheet5: 统计汇总
    total     = len(rows)
    rule_err  = sum(1 for r in rows if r.get("是否错误")   == "❗错误")
    ai_err    = sum(1 for r in rows if r.get("AI是否正确") == "❗错误") if has_ai else 0
    consistent = sum(1 for r in rows if r.get("一致性")    == "✅一致")

    stats = pd.DataFrame([
        ["总行数",       total],
        ["规则检出错误", rule_err],
        ["AI检出错误",   ai_err],
        ["确认错误",     len(confirmed_rows)],
        ["漏检（FN）",   len(fn_rows)],
        ["误报（FP）",   len(fp_rows)],
        ["规则+AI一致",  consistent],
    ], columns=["指标", "数量"])
    _write_df(wb.create_sheet("统计汇总"), stats, "4472C4", [20, 10])

    wb.save(output_path)
    print(f"✅ 合并报告: {output_path}")
    return output_path


# ─────────────────────────────────────────
# 2. 术语一致性检查报告（原有，保持兼容）
# ─────────────────────────────────────────

def generate_report(check_result: Dict,
                    glossary: List[Dict],
                    output_path: str = "reports/术语一致性检查报告.xlsx"):
    _ensure_dir(output_path)
    if output_path == "reports/术语一致性检查报告.xlsx":
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_path.replace(".xlsx", f"_{ts}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)

    def sh(ws, color):
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

    freq = check_result.get("frequency", {})
    if freq:
        ws1 = wb.create_sheet("术语频率统计")
        ws1.append(["原文术语", "标准译法", "原文句匹配数", "译文句匹配数"])
        sh(ws1, "4169E1")
        for term, info in sorted(freq.items(), key=lambda x: x[1]["count"], reverse=True):
            ws1.append([term, info["standard_translation"], info["count"], info.get("target_count", 0)])
        for col, w in [("A", 25), ("B", 30), ("C", 18), ("D", 18)]:
            ws1.column_dimensions[col].width = w

    found = check_result.get("found_terms", [])
    if found:
        ws2 = wb.create_sheet("程序检查结果")
        ws2.append(["原文术语", "标准译法", "实际译法", "是否合规", "原文句子", "译文句子", "行号"])
        sh(ws2, "708090")
        for r in found:
            ws2.append([r.get("term"), r.get("standard_translation"), r.get("actual_translation"),
                        "✓" if r.get("is_compliant") else "✗",
                        r.get("source_sentence"), r.get("target_sentence"), r.get("row")])
        for col, w in [("A", 20), ("B", 25), ("C", 25), ("D", 10), ("E", 50), ("F", 50), ("G", 8)]:
            ws2.column_dimensions[col].width = w

    violations = check_result.get("violations", [])
    if violations:
        ws3 = wb.create_sheet("LLM检查结果")
        ws3.append(["原文术语", "标准译法", "实际译法（LLM）", "相似度", "原文句子", "译文句子", "行号"])
        sh(ws3, "4B0082")
        for v in violations:
            sim = v.get("similarity", "")
            ws3.append([v.get("term"), v.get("standard_translation"), v.get("_revision_actual"),
                        f"{sim}%", v.get("source_sentence"), v.get("target_sentence"), v.get("row")])
        for col, w in [("A", 20), ("B", 25), ("C", 25), ("D", 12), ("E", 50), ("F", 50), ("G", 8)]:
            ws3.column_dimensions[col].width = w

    if found:
        ws4 = wb.create_sheet("合并检查结果")
        ws4.append(["原文术语", "标准译法", "实际译法", "是否合规", "相似度", "原文句子", "译文句子", "行号"])
        sh(ws4, "DC143C")
        for r in found:
            actual = r.get("actual_translation") if r.get("is_compliant") else (r.get("_revision_actual") or r.get("actual_translation"))
            sim = "100%" if r.get("is_compliant") else f"{r.get('similarity', '')}%"
            ws4.append([r.get("term"), r.get("standard_translation"), actual,
                        "✓" if r.get("is_compliant") else "✗", sim,
                        r.get("source_sentence"), r.get("target_sentence"), r.get("row")])
        for col, w in [("A", 20), ("B", 25), ("C", 25), ("D", 10), ("E", 12), ("F", 50), ("G", 50), ("H", 8)]:
            ws4.column_dimensions[col].width = w

    wb.save(output_path)
    print(f"✅ 术语报告: {output_path}")
    return output_path


# ─────────────────────────────────────────
# 3. 对照对齐报告
# ─────────────────────────────────────────

def generate_align_report(rows: List[Dict],
                           output_path: str = "reports/final_align.xlsx"):
    """
    接收 对照.run() 返回的 List[Dict]，生成对齐 Excel。

    每项期望字段：chunk_id, source, target, hash
    Sheet：对齐结果（序号 / 原文 / 译文 / chunk_id）
    """
    _ensure_dir(output_path)
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("对齐结果")
    ws.append(["序号", "原文", "译文", "chunk_id"])
    _style_header(ws, "2E75B6")
    _set_col_widths(ws, [8, 70, 70, 10])

    for i, item in enumerate(rows, 1):
        ws.append([i, item.get("source", ""), item.get("target", ""), item.get("chunk_id", "")])

    wb.save(output_path)
    print(f"✅ 对齐报告: {output_path}")
    return output_path
