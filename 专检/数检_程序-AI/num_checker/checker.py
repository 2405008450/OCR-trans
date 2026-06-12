"""
checker.py — 主检查流程入口
============================
整合模块A（symbolic_parser）和模块B（rl_discriminator），
对原文/译文句对进行纯程序数值检查。

流程：
  原文 → [模块A] 提取数值 → [模块B] 多义词过滤 → 规范化数值列表
  译文 → [模块A] 提取数值 → [模块B] 多义词过滤 → 规范化数值列表
  两侧数值列表 → [对齐矩阵] → 差异报告
"""

import os
from dataclasses import dataclass, field
from typing import List, Tuple

from .symbolic_parser import parse_values
from .crf_discriminator import filter_ambiguous_tokens  # CRF优先，无模型自动回退到规则
from .alignment_matrix import build_matrix, format_errors, AlignError


# ─────────────────────────────────────────
# 检查结果数据类
# ─────────────────────────────────────────

@dataclass
class CheckResult:
    src_text:   str
    tgt_text:   str
    src_values: List[str]
    tgt_values: List[str]
    errors:     List[AlignError]
    is_correct: bool
    summary:    str

    def to_dict(self) -> dict:
        return {
            "原文":     self.src_text,
            "译文":     self.tgt_text,
            "原文数值": self.src_values,
            "译文数值": self.tgt_values,
            "检查结果": "错误" if not self.is_correct else "正确",
            "错误类型": "；".join(e.error_type for e in self.errors),
            "错误原因": "；".join(e.message for e in self.errors),
        }


# ─────────────────────────────────────────
# 核心检查函数
# ─────────────────────────────────────────

def check_pair(src: str, tgt: str, use_rl_filter: bool = True) -> CheckResult:
    """对单个原文/译文句对进行数值检查"""
    src_values = parse_values(src)
    tgt_values = parse_values(tgt)

    if use_rl_filter:
        src_values = filter_ambiguous_tokens(src, src_values)
        tgt_values = filter_ambiguous_tokens(tgt, tgt_values)

    errors = build_matrix(src_values, tgt_values)
    return CheckResult(
        src_text=src, tgt_text=tgt,
        src_values=src_values, tgt_values=tgt_values,
        errors=errors, is_correct=len(errors) == 0,
        summary=format_errors(errors, src, tgt),
    )


def check_pairs(pairs: List[Tuple[str, str]], use_rl_filter: bool = True) -> List[CheckResult]:
    return [check_pair(src, tgt, use_rl_filter) for src, tgt in pairs]


def check_excel(alignment_path: str, use_rl_filter: bool = True) -> List[CheckResult]:
    """从对照 Excel（含原文/译文列）读取并检查"""
    import pandas as pd
    df = pd.read_excel(alignment_path)
    pairs = [(str(row.get("原文", "")), str(row.get("译文", ""))) for _, row in df.iterrows()]
    return check_pairs(pairs, use_rl_filter)


# ─────────────────────────────────────────
# 报告生成
# ─────────────────────────────────────────

def save_report(results: List[CheckResult], output_path: str, include_correct: bool = True):
    import pandas as pd
    from openpyxl.styles import PatternFill

    rows = [r.to_dict() for r in results]  # 始终输出全部行
    if not rows:
        rows = [{"原文":"","译文":"","原文数值":[],"译文数值":[],"检查结果":"正确","错误类型":"","错误原因":"无数据"}]

    df = pd.DataFrame(rows)
    for col in ["原文数值", "译文数值"]:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: "，".join(x) if isinstance(x, list) else str(x))

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="数值检查")
        ws = writer.sheets["数值检查"]
        red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        for row_idx, row in enumerate(df.itertuples(), start=2):
            if getattr(row, "检查结果", "") == "错误":
                for cell in ws[row_idx]:
                    cell.fill = red
        for col in ws.columns:
            w = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(w + 4, 60)

    total  = len(results)
    errors = sum(1 for r in results if not r.is_correct)
    print(f"报告已保存: {output_path}")
    print(f"   总计: {total} 行  |  错误: {errors} 行  |  正确率: {(total-errors)/total*100:.1f}%")


def print_summary(results: List[CheckResult]):
    total  = len(results)
    errors = [r for r in results if not r.is_correct]
    print(f"\n{'='*60}")
    print("纯程序数值检查结果")
    print(f"{'='*60}")
    print(f"  总行数: {total}  |  错误: {len(errors)}  |  正确率: {(total-len(errors))/total*100:.1f}%" if total else "  无数据")
    if errors:
        print(f"\n{'─'*60}")
        for i, r in enumerate(errors[:20], 1):
            print(f"\n  [{i}] 原文: {r.src_text[:80]}")
            print(f"       译文: {r.tgt_text[:80]}")
            print(f"       原文数值: {r.src_values}")
            print(f"       译文数值: {r.tgt_values}")
            for e in r.errors:
                print(f"       [!] {e.message}")
    print(f"{'='*60}\n")
