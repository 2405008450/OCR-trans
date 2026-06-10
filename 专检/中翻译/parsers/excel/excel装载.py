import traceback
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
import json
from openpyxl.utils import get_column_letter
from parsers.json.clean_json import extract_and_parse


class ExcelReportGenerator:
    def __init__(self, output_dir="reports"):
        self.output_dir = output_dir
        if not os.path.exists(self.output_dir): os.makedirs(self.output_dir)

    def generate_report(self, filename_prefix, sheets_config):
        wb = Workbook()
        wb.remove(wb.active)
        for config in sheets_config:
            ws = wb.create_sheet(config["name"])
            ws.append(config["header"])
            for row in config["data"]: ws.append(row)
            # 简单样式
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
            wb.save(os.path.join(self.output_dir, f"{filename_prefix}_{datetime.now().strftime('%m%d_%H%M')}.xlsx"))

    @staticmethod
    def get_excel(data_or_path, output_dir=None):
        """
        解析并导出到指定的 Excel 文件夹
        :param data_or_path: 输入的 JSON 结果路径（字符串）或已解析的列表数据
        :param output_dir: 指定存放 Excel 的文件夹路径
        :return: (filename, df)
        """
        # 判断输入类型
        if isinstance(data_or_path, str):
            # 如果是字符串，当作文件路径处理
            results = extract_and_parse(data_or_path)
        elif isinstance(data_or_path, list):
            # 如果是列表，直接使用
            results = data_or_path
        else:
            print(f"❌ 不支持的输入类型: {type(data_or_path)}")
            return None, None

        if not results:
            print("⚠️ 未提取到有效数据。")
            return None, None

        # --- 新增：处理文件夹路径逻辑 ---
        if output_dir:
            output_path = Path(output_dir)
            # 如果文件夹不存在，自动递归创建
            output_path.mkdir(parents=True, exist_ok=True)
        else:
            # 默认当前脚本运行目录
            output_path = Path(".")

        try:
            term_data = results
            print(f"\n🚀 正在生成 Excel 报告...")
            wb = Workbook()
            ws = wb.active
            ws.title = "质检报告"

            # 设置表头
            headers = ["错误编号", "错误类型", "原文数值", "译文数值", "译文建议修改值", "修改理由", "违反的规则",
                       "原文上下文", "译文上下文", "原文位置", "译文位置", "替换锚点"]
            ws.append(headers)

            # 美化表头风格
            header_fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 填充数据
            for item in term_data:
                ws.append([
                    item.get("错误编号", ""), item.get("错误类型", ""),
                    item.get("原文数值", ""), item.get("译文数值", ""),
                    item.get("译文修改建议值", ""), item.get("修改理由", ""),
                    item.get("违反的规则", ""), item.get("原文上下文", ""),
                    item.get("译文上下文", ""), item.get("原文位置", ""),
                    item.get("译文位置", ""), item.get("替换锚点", "")
                ])

            # 自动调整列宽
            column_widths = [8, 15, 20, 20, 20, 30, 40, 40, 40, 10, 10, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # --- 关键修改：合成完整的文件保存路径 ---
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_name = f"数值质检报告_{timestamp}.xlsx"
            full_save_path = output_path / file_name  # 使用 pathlib 拼接路径

            # 保存到指定位置
            wb.save(str(full_save_path))
            print(f"✅ 导出成功: {full_save_path}")

            # 重新读取刚才保存的文件
            try:
                df = pd.read_excel(str(full_save_path), engine='openpyxl')
                # 返回 完整路径字符串 和 数据对象
                return str(full_save_path), df
            except Exception as read_e:
                print(f"❌ 读取已保存文件失败: {read_e}")
                return str(full_save_path), pd.DataFrame(results)

        except Exception as e:
            print(f"❌ 发生错误: {e}")
            traceback.print_exc()
            return None, None

if __name__ == '__main__':
    # --- 新增：解析并导出到 Excel ---

    body_result_path = r"C:\Users\Administrator\Desktop\中翻译通用规则项目\zhongfanyi\llm\llm_project\zhengwen\output_json\文本对比结果_20260312_134420.json"
    # e=ExcelReportGenerator()
    # file_path,data=e.get_excel(body_result_path)
    # print(f"报告已存至: {os.path.abspath(file_path)}")
    # print(data)
    # 1. 比如你想存放在桌面下的 'Reports' 文件夹
    target_folder = r"C:\Users\Administrator\Desktop\中翻译通用规则项目\zhongfanyi\llm\llm_project\yejiao\output_excel"
    e = ExcelReportGenerator()
    # 调用时传入文件夹路径
    save_path, data = e.get_excel(body_result_path, output_dir=target_folder)
    if save_path:
        print(f"文件成功保存在：{save_path}")
        print(data)

    # print("\n正在启动超强提取模式...")
    # results = extract_and_parse(body_result_path)
    # print(results)
    # for res in results:
    #     print(res)
    #
    # if results:
    #     print(f"✅ 解析成功！共获取 {len(results)} 个错误条目。")
    #     print(f"第一个错误: ID {results[0].get('错误编号')} - {results[0].get('错误类型')}")
    # else:
    #     print("❌ 依然没有提取到数据。请再次运行并告诉我报错信息。")
    #
    # # --- 新增：解析并导出到 Excel ---
    # try:
    #     # 将大模型返回的字符串解析为 Python 列表
    #     term_data = results
    #     print(term_data)
    #
    #     if term_data:
    #         print(f"\n🚀 正在生成 Excel 报告...")
    #         wb = Workbook()
    #         ws = wb.active
    #         ws.title = "潜在术语提取"
    #
    #         # 设置表头 (匹配你的需求)
    #         headers = ["错误编号", "错误类型", "原文数值","译文数值", "译文建议修改值", "修改理由","违反的规则","原文上下文","译文上下文","原文位置","译文位置","替换锚点"]
    #         ws.append(headers)
    #
    #         # 美化表头
    #         header_fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
    #         header_font = Font(bold=True, color="FFFFFF")
    #         for cell in ws[1]:
    #             cell.fill = header_fill
    #             cell.font = header_font
    #             cell.alignment = Alignment(horizontal="center", vertical="center")
    #
    #         # 填充数据：将 JSON 字段映射到 Excel 列
    #         for item in term_data:
    #             ws.append([
    #                 item.get("错误编号", ""),
    #                 item.get("错误类型", ""),
    #                 item.get("原文数值", ""),
    #                 item.get("译文数值", ""),
    #                 item.get("译文修改建议值", ""),
    #                 item.get("修改理由", ""),
    #                 item.get("违反的规则", ""),
    #                 item.get("原文上下文", ""),
    #                 item.get("译文上下文", ""),
    #                 item.get("原文位置", ""),
    #                 item.get("译文位置", ""),
    #                 item.get("替换锚点", "")
    #             ])
    #
    #         # 4. 调整列宽以适应内容
    #         column_widths = [8, 15, 20, 20, 20, 30, 40, 40, 40, 10, 10, 20]
    #         for i, width in enumerate(column_widths, 1):
    #             ws.column_dimensions[get_column_letter(i)].width = width
    #
    #         # 保存文件
    #         timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    #         filename = f"潜在术语提取_{timestamp}.xlsx"
    #         wb.save(filename)
    #         print(f"✅ 导出成功: {filename}")
    #     else:
    #         print("⚠️ 未提取到有效术语。")
    #
    # except json.JSONDecodeError as e:
    #     print(f"❌ JSON 解析失败，请检查模型输出格式: {e}")
    #     print("原始输出内容：", results)
    # except Exception as e:
    #     print(f"❌ 发生错误: {e}")
    #     traceback.print_exc()