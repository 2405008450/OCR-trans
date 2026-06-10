# -*- coding: utf-8 -*-
"""
word_doc_full_extractor.py
从Word文档的9个位置提取所有文字：正文、表格、文本框、图表、页眉、页脚、脚注、尾注、批注
接口函数：extract_all_text(doc_path) -> str

优化版：只打开ZIP一次，全部用lxml解析，避免python-docx的开销。
"""
from lxml import etree
from zipfile import ZipFile
from datetime import datetime
import os, time

W_NS  = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'
A_NS  = 'http://schemas.openxmlformats.org/drawingml/2006/main'
C_NS  = 'http://schemas.openxmlformats.org/drawingml/2006/chart'
WPS_NS = 'http://schemas.microsoft.com/office/word/2010/wordprocessingShape'

W   = '{' + W_NS  + '}'
A   = '{' + A_NS  + '}'
C   = '{' + C_NS  + '}'
WPS = '{' + WPS_NS + '}'

# ── Excel输出路径 ──
EXCEL_OUT_DIR = os.path.dirname(os.path.abspath(__file__))


def _para_texts(root):
    """从XML root中提取所有 <w:p> 段落的纯文本，返回非空文本列表"""
    texts = []
    for p in root.iter(f'{W}p'):
        parts = []
        for t in p.iter(f'{W}t'):
            if t.text:
                parts.append(t.text)
        line = ''.join(parts).strip()
        if line:
            texts.append(line)
    return texts


def _extract_all_from_zip(doc_path):
    """一次性打开ZIP，提取9个位置的所有文字"""
    body_texts     = []
    table_texts    = []  # 正文和表格在document.xml中一起提取
    textbox_texts  = []
    header_texts   = []
    footer_texts   = []
    footnote_texts = []
    endnote_texts  = []
    comment_texts  = []
    chart_texts    = []

    with ZipFile(doc_path, 'r') as zf:
        names = set(zf.namelist())

        # ── 1 & 2 & 3: 正文 + 表格 + 文本框（都在 document.xml 中）──
        if 'word/document.xml' in names:
            doc_root = etree.fromstring(zf.read('word/document.xml'))

            # 文本框：先提取并记录其元素，后面正文提取时跳过
            txbx_elements = set()
            for tc in doc_root.iter(f'{WPS}txbxContent'):
                txbx_elements.add(id(tc))
                textbox_texts.extend(_para_texts(tc))
            # 兼容其他命名空间的文本框
            if not textbox_texts:
                for el in doc_root.iter():
                    if el.tag.endswith('txbxContent'):
                        txbx_elements.add(id(el))
                        textbox_texts.extend(_para_texts(el))

            # 正文段落（w:body 直接子 w:p，排除表格和文本框内的）
            body = doc_root.find(f'{W}body')
            if body is not None:
                for child in body:
                    if child.tag == f'{W}p':
                        # 检查是否在文本框内（通过祖先判断——顶层段落不会在文本框内）
                        parts = []
                        for t in child.iter(f'{W}t'):
                            if t.text:
                                parts.append(t.text)
                        line = ''.join(parts).strip()
                        if line:
                            body_texts.append(line)
                    elif child.tag == f'{W}tbl':
                        # 表格
                        table_texts.extend(_para_texts(child))
                    elif child.tag == f'{W}sdt':
                        # 结构化文档标签（可能包含段落和表格）
                        for p in child.iter(f'{W}p'):
                            parts = []
                            for t in p.iter(f'{W}t'):
                                if t.text:
                                    parts.append(t.text)
                            line = ''.join(parts).strip()
                            if line:
                                body_texts.append(line)

        # ── 4: 图表 ──
        chart_files = [f for f in names
                       if f.startswith('word/charts/') and f.endswith('.xml')]
        for cf in chart_files:
            root = etree.fromstring(zf.read(cf))
            seen = set()
            # DrawingML <a:t>
            for t_elem in root.iter(f'{A}t'):
                t = (t_elem.text or '').strip()
                if t and t not in seen:
                    seen.add(t)
                    chart_texts.append(t)
            # Chart <c:v> 字符串值
            for v_elem in root.iter(f'{C}v'):
                t = (v_elem.text or '').strip()
                if t and t not in seen:
                    try:
                        float(t)
                    except ValueError:
                        seen.add(t)
                        chart_texts.append(t)
            # Chart <c:tx> 文本引用
            for tx_elem in root.iter(f'{C}tx'):
                for sv in tx_elem.iter(f'{C}v'):
                    t = (sv.text or '').strip()
                    if t and t not in seen:
                        seen.add(t)
                        chart_texts.append(t)

        # ── 5 & 6: 页眉 & 页脚 ──
        for fname in names:
            if fname.startswith('word/header') and fname.endswith('.xml'):
                root = etree.fromstring(zf.read(fname))
                header_texts.extend(_para_texts(root))
            elif fname.startswith('word/footer') and fname.endswith('.xml'):
                root = etree.fromstring(zf.read(fname))
                footer_texts.extend(_para_texts(root))

        # ── 7: 脚注 ──
        if 'word/footnotes.xml' in names:
            root = etree.fromstring(zf.read('word/footnotes.xml'))
            for fn in root.iter(f'{W}footnote'):
                ftype = fn.get(f'{W}type')
                if ftype in ('separator', 'continuationSeparator'):
                    continue
                footnote_texts.extend(_para_texts(fn))

        # ── 8: 尾注 ──
        if 'word/endnotes.xml' in names:
            root = etree.fromstring(zf.read('word/endnotes.xml'))
            for en in root.iter(f'{W}endnote'):
                etype = en.get(f'{W}type')
                if etype in ('separator', 'continuationSeparator'):
                    continue
                endnote_texts.extend(_para_texts(en))

        # ── 9: 批注 ──
        if 'word/comments.xml' in names:
            root = etree.fromstring(zf.read('word/comments.xml'))
            for c in root.iter(f'{W}comment'):
                comment_texts.extend(_para_texts(c))

    return [
        ("正文", body_texts),
        ("表格", table_texts),
        ("文本框", textbox_texts),
        ("图表", chart_texts),
        ("页眉", header_texts),
        ("页脚", footer_texts),
        ("脚注", footnote_texts),
        ("尾注", endnote_texts),
        ("批注", comment_texts),
    ]


# ══════════════════════════════════════════
# 保存Excel记录
# ══════════════════════════════════════════

def _save_extraction_excel(results, start_dt, end_dt):
    """将各位置提取的文字保存到Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "提取内容"

    headers = ["序号", "位置", "提取文字"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    seq = 1
    for loc, texts in results:
        for t in texts:
            ws.cell(row=row_idx, column=1, value=seq)
            ws.cell(row=row_idx, column=2, value=loc)
            ws.cell(row=row_idx, column=3, value=t)
            row_idx += 1
            seq += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 100

    time_tag = f"{start_dt.strftime('%y%m%d %H%M')}-{end_dt.strftime('%H%M')}"
    filename = f"word_doc_full_extractor 测试1-1-1 {time_tag}.xlsx"
    os.makedirs(EXCEL_OUT_DIR, exist_ok=True)
    xlsx_path = os.path.join(EXCEL_OUT_DIR, filename)
    wb.save(xlsx_path)
    print(f"[word_doc_full_extractor] Excel已保存: {xlsx_path}")
    return xlsx_path


# ══════════════════════════════════════════
# 主接口函数
# ══════════════════════════════════════════

def extract_all_text(doc_path):
    """
    从Word文档的9个位置提取所有文字，返回合并后的纯文本字符串。
    9个位置：正文、表格、文本框、图表、页眉、页脚、脚注、尾注、批注

    参数:
        doc_path: Word文档路径 (.docx)
    返回:
        str: 所有位置的文字合并后的字符串
    """
    start_dt = datetime.now()
    t0 = time.time()
    print(f"[word_doc_full_extractor] 开始提取: {doc_path}")

    results = _extract_all_from_zip(doc_path)

    for loc, texts in results:
        print(f"    {loc}: {len(texts)} 条")

    total_items = sum(len(t) for _, t in results)
    print(f"[word_doc_full_extractor] 共提取 {total_items} 条文字, 耗时: {time.time()-t0:.2f}s")

    # 保存Excel记录
    end_dt = datetime.now()
    try:
        _save_extraction_excel(results, start_dt, end_dt)
    except Exception as e:
        print(f"[word_doc_full_extractor] 保存Excel时出错: {e}")

    # 返回合并后的纯文本
    all_texts = []
    for _, texts in results:
        all_texts.extend(texts)
    full_text = "\n".join(all_texts)
    print(f"[word_doc_full_extractor] 返回文本共 {len(full_text)} 字符")
    return full_text


# ── 独立运行测试 ──
if __name__ == "__main__":
    # import sys
    # if len(sys.argv) > 1:
    #     path = sys.argv[1]
    # else:
    #     print("用法: python word_doc_full_extractor.py <docx文件路径>")
    #     sys.exit(1)
    path=r"C:\Users\H\Desktop\测试项目\中翻译\测试文件\雅本化学2025ESG报告文字稿-20260409.docx"
    text = extract_all_text(path)
    print(f"\n提取完成，共 {len(text)} 字符")
    print(text)
