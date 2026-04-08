import openpyxl
import pandas as pd


def _get_cell_value_dedup(ws, row, col, merged_origins):
    """
    获取单元格的值，合并单元格只在左上角输出一次。
    merged_origins: set of (min_row, min_col) 已记录的合并区域起点。
    """
    cell = ws.cell(row=row, column=col)

    # 普通单元格，直接返回
    if cell.value is not None:
        # 检查是否恰好是合并区域的左上角（有值的那个）
        for merged_range in ws.merged_cells.ranges:
            if row == merged_range.min_row and col == merged_range.min_col:
                merged_origins.add((row, col))
                return cell.value
        # 不在任何合并区域内，普通单元格
        return cell.value

    # 值为 None，检查是否在合并区域内（非左上角的部分）
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # 是合并区域的非起始单元格，跳过
            return None
    return None


def parse_excel_with_openpyxl(file_path):
    """
    使用 openpyxl 解析 Excel，正确处理合并单元格。
    支持 .xlsx 格式。
    返回所有 sheet 的文本内容拼接字符串。
    """
    try:
        wb = openpyxl.load_workbook(str(file_path), data_only=True)
        all_text_parts = []

        for ws in wb.worksheets:
            sheet_lines = []
            merged_origins = set()
            for row_idx in range(1, ws.max_row + 1):
                row_texts = []
                for col_idx in range(1, ws.max_column + 1):
                    val = _get_cell_value_dedup(ws, row_idx, col_idx, merged_origins)
                    if val is not None:
                        text = str(val).strip()
                        if text:
                            row_texts.append(text)
                if row_texts:
                    sheet_lines.append("\t".join(row_texts))

            if sheet_lines:
                all_text_parts.append(f"[{ws.title}]\n" + "\n".join(sheet_lines))

        wb.close()
        return "\n\n".join(all_text_parts)

    except Exception as e:
        print(f"openpyxl 解析出错: {e}")
        return ""


def parse_excel_with_pandas(file_path):
    """
    解析 Excel 文件，支持多个 sheet。
    .xlsx 优先用 openpyxl 直接解析（正确处理合并单元格）。
    .xls 用 pandas + xlrd 引擎。
    返回所有 sheet 的文本内容拼接字符串。
    """
    file_str = str(file_path).lower()

    # .xlsx 用 openpyxl 直接解析，合并单元格处理更好
    if file_str.endswith(".xlsx"):
        result = parse_excel_with_openpyxl(file_path)
        if result:
            return result

    # .xls 或 openpyxl 失败时回退到 pandas
    try:
        engine = "xlrd" if file_str.endswith(".xls") else None
        all_sheets = pd.read_excel(file_path, sheet_name=None, engine=engine, header=None)

        all_text_parts = []
        for sheet_name, df in all_sheets.items():
            # 去掉全为 NaN 的行和列
            df = df.dropna(how="all").dropna(axis=1, how="all")
            sheet_lines = []
            for _, row in df.iterrows():
                row_texts = []
                for val in row:
                    if pd.notna(val):
                        text = str(val).strip()
                        if text:
                            row_texts.append(text)
                if row_texts:
                    sheet_lines.append("\t".join(row_texts))
            if sheet_lines:
                all_text_parts.append(f"[{sheet_name}]\n" + "\n".join(sheet_lines))

        return "\n\n".join(all_text_parts)

    except Exception as e:
        print(f"解析出错: {e}")
        return ""
if __name__ == '__main__':
    path=r"C:\Users\H\Desktop\测试项目\中翻译\测试文件\原文-副本ChatGPT差异化AI内容生成模板.xlsx"
    r=parse_excel_with_pandas(path)
    print(r)