import json
import re
import codecs
import traceback
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from openpyxl.utils import get_column_letter

def extract_and_parse(file_path: str) -> List[dict]:
    # 1. 读取原始内容
    try:
        with codecs.open(file_path, 'r', encoding='utf-8-sig') as f:
            raw_content = f.read()
    except Exception as e:
        print(f"❌ 读取文件失败: {e}")
        return []

    return parse_json_content(raw_content)


def parse_json_content(raw_content: str) -> List[dict]:
    """从字符串中解析 JSON 错误列表。支持 Markdown 代码块、转义字符串等格式。"""
    if not raw_content or not raw_content.strip():
        return []

    # 2. 【核心修复】解包逻辑
    # 如果文件内容看起来像是一个被转义的字符串 (以 " 开头)，先把它还原
    content = raw_content
    if content.strip().startswith('"'):
        try:
            # json.loads 会处理掉外层的引号，并将 \\n 还原为真正的换行符
            content = json.loads(content)
            print("✅ 成功解包：检测到文件是序列化字符串，已完成解包。")
        except json.JSONDecodeError:
            print("⚠️ 尝试解包失败，将按原样处理。")

    # 2.5 移除 Markdown 代码块标记
    if content.strip().startswith('```'):
        lines = content.split('\n')
        # 找到第一个 ``` 和最后一个 ```
        start_idx = -1
        end_idx = -1
        for i, line in enumerate(lines):
            if line.strip().startswith('```'):
                if start_idx == -1:
                    start_idx = i
                else:
                    end_idx = i
        
        if start_idx != -1 and end_idx != -1:
            json_lines = lines[start_idx + 1:end_idx]
            content = '\n'.join(json_lines)
            print("✅ 移除 Markdown 代码块标记")

    # 3. 尝试直接解析 JSON（如果格式正确）
    try:
        data = json.loads(content)
        if isinstance(data, list):
            print(f"✅ 直接解析成功：提取到 {len(data)} 个对象")
            # 排序
            data.sort(key=lambda x: int(x.get("错误编号", 99999)))
            return data
    except json.JSONDecodeError as e:
        print(f"⚠️ 直接解析失败 ({e.msg})，使用栈式提取...")

    # 4. 栈式提取器：从 Markdown 中提取所有 {} 块
    # 【关键修复】在提取每个对象后，修复控制字符问题
    objects = []
    stack_depth = 0
    start_index = -1

    for i, char in enumerate(content):
        if char == '{':
            if stack_depth == 0:
                start_index = i
            stack_depth += 1
        elif char == '}':
            if stack_depth > 0:
                stack_depth -= 1
                if stack_depth == 0 and start_index != -1:
                    json_str = content[start_index: i + 1]
                    
                    # 【关键修复】尝试修复控制字符问题
                    # 方法1: 直接解析
                    try:
                        obj = json.loads(json_str)
                        objects.append(obj)
                    except json.JSONDecodeError:
                        # 方法2: 替换未转义的控制字符后再解析
                        try:
                            # 使用 repr 和 eval 来处理控制字符
                            # 这会自动转义所有控制字符
                            fixed_str = json_str.encode('unicode_escape').decode('ascii')
                            # 但这会双重转义已经转义的字符，所以需要还原
                            fixed_str = fixed_str.replace('\\\\', '\\')
                            obj = json.loads(fixed_str)
                            objects.append(obj)
                            print(f"  ✓ 修复控制字符后成功解析错误 {obj.get('错误编号', '?')}")
                        except:
                            # 方法3: 使用 strict=False（Python 3.6+）
                            try:
                                # 先手动替换常见的控制字符
                                fixed_str = json_str
                                # 不要替换已经转义的 \t \n 等
                                # 只替换未转义的制表符和换行符
                                import re
                                # 查找未转义的制表符（不在 \" 之间的 \t）
                                # 简化处理：直接替换所有制表符为 \\t
                                fixed_str = fixed_str.replace('\t', '\\t')
                                fixed_str = fixed_str.replace('\r', '\\r')
                                # 但保留 JSON 结构中的换行符
                                
                                obj = json.loads(fixed_str)
                                objects.append(obj)
                                print(f"  ✓ 手动修复控制字符后成功解析错误 {obj.get('错误编号', '?')}")
                            except Exception as final_error:
                                print(f"  ✗ 无法解析对象 (位置 {start_index}-{i}): {final_error}")
                    
                    start_index = -1

    # 5. 排序
    objects.sort(key=lambda x: int(x.get("错误编号", 99999)))

    print(f"DEBUG: 成功识别并提取了 {len(objects)} 个完整对象。")
    return objects


if __name__ == "__main__":

    body_result_path = r"C:\Users\Administrator\Desktop\中翻译通用规则项目\zhongfanyi\llm\llm_project\zhengwen\output_json\文本对比结果_20260312_134420.json"
    print("\n正在启动超强提取模式...")
    results = extract_and_parse(body_result_path)
    print(results)
    for res in results:
        print(res)

    if results:
        print(f"✅ 解析成功！共获取 {len(results)} 个错误条目。")
        print(f"第一个错误: ID {results[0].get('错误编号')} - {results[0].get('错误类型')}")
    else:
        print("❌ 依然没有提取到数据。请再次运行并告诉我报错信息。")

    # --- 新增：解析并导出到 Excel ---
    try:
        # 将大模型返回的字符串解析为 Python 列表
        term_data = results
        print(term_data)

        if term_data:
            print(f"\n🚀 正在生成 Excel 报告...")
            wb = Workbook()
            ws = wb.active
            ws.title = "潜在术语提取"

            # 设置表头 (匹配你的需求)
            headers = ["错误编号", "错误类型", "原文数值","译文数值", "译文建议修改值", "修改理由","违反的规则","原文上下文","译文上下文","原文位置","译文位置","替换锚点"]
            ws.append(headers)

            # 美化表头
            header_fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 填充数据：将 JSON 字段映射到 Excel 列
            for item in term_data:
                ws.append([
                    item.get("错误编号", ""),
                    item.get("错误类型", ""),
                    item.get("原文数值", ""),
                    item.get("译文数值", ""),
                    item.get("译文修改建议值", ""),
                    item.get("修改理由", ""),
                    item.get("违反的规则", ""),
                    item.get("原文上下文", ""),
                    item.get("译文上下文", ""),
                    item.get("原文位置", ""),
                    item.get("译文位置", ""),
                    item.get("替换锚点", "")
                ])

            # 4. 调整列宽以适应内容
            column_widths = [8, 15, 20, 20, 20, 30, 40, 40, 40, 10, 10, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # 保存文件
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"潜在术语提取_{timestamp}.xlsx"
            wb.save(filename)
            print(f"✅ 导出成功: {filename}")
        else:
            print("⚠️ 未提取到有效术语。")

    except json.JSONDecodeError as e:
        print(f"❌ JSON 解析失败，请检查模型输出格式: {e}")
        print("原始输出内容：", results)
    except Exception as e:
        print(f"❌ 发生错误: {e}")
        traceback.print_exc()