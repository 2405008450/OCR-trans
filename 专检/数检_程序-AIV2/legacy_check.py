"""
旧版流程适配（Mode B · DOCX · 无程序对齐）

对应 old_check/old_number/main.py 的旧逻辑：
  1. 整篇提取原文/译文（正文/页眉/页脚三通道，不做逐段配对）
  2. 按字数分块（text_splitter.split_text_pair，带缓冲区重叠）
  3. 每块直接送 AI 对比（无规则预检），AI 独立判断整块内的所有数值问题
  4. 合并去重为扁平错误列表（不含 para_index/行号，仅靠锚点+上下文定位）
  5. 写回时仅用锚点/上下文匹配（降级但更贴合旧流程）

与当前 Mode B（_build_pairs_from_docx 等）的区别：不做原文/译文逐段对齐，
避免"结构不符"类错误在长文档、格式差异较大的原文/译文之间频繁出现。
"""
import os
import time
import json
from typing import List, Dict, Tuple

from dotenv import load_dotenv
from openai import OpenAI

from full_content import scan_docx
from header_extractor import extract_headers
from footer_extractor import extract_footers
from text_splitter import split_text_pair, split_bilingual_text, _count_chars
from excel.clean_json import parse_json_content
from clean_replace_duplicates import clean_suggestion

load_dotenv()
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"), base_url=os.getenv("BASE_URL"))

_MODEL = "deepseek/deepseek-v4-pro"

# 正文相关 source（与 main.py 的 _BODY_SOURCES 保持一致，排除 header/footer 避免重复）
_BODY_SOURCES = {"body", "table", "toc", "footnote", "endnote", "chart", "textbox"}


# =============================================================
# 整篇文本提取（三通道：正文/页眉/页脚）
# =============================================================

def _extract_body_only(docx_path: str) -> str:
    """提取正文相关内容（不含页眉页脚），按阅读顺序拼接为整篇文本。"""
    segs = scan_docx(docx_path)
    lines = [s.text for s in segs if s.source in _BODY_SOURCES and s.text and s.text.strip()]
    return "\n".join(lines)


def extract_channels(docx_path: str) -> Tuple[str, str, str]:
    """返回 (正文, 页眉, 页脚) 三个整篇文本。"""
    body = _extract_body_only(docx_path)
    headers = extract_headers(docx_path)
    footers = extract_footers(docx_path)
    return body, "\n".join(headers), "\n".join(footers)


# =============================================================
# LLM 对比（整块直接送检，无规则预检）
# =============================================================

_PROMPT_TEMPLATE = """# 角色
你是一名资深的文件翻译审校专家，精通中英文以及其他语种文件中的数值、金额、日期、编号层级的一致性核查。

# 任务
1. 严格对比【原文】与【译文】，识别并提取所有数值、单位、编号、日期翻译后译文与原文不一致的错误。
2. 对照原文检查译文出现错译、漏译、多译的情况，以【原文】为唯一准则。
3. 若译文符合规则且无错译，严禁将其放入 JSON 列表中输出。

# 审查规则（执行优先级最高）：
1. 原文主权原则：修改建议必须以【原文】为唯一准则，禁止引入原文中不存在的信息。
2. 提取纯粹性：严禁拼接、计算或修改数字。仅当译文数值与原文在含义、精确度或逻辑上不符时才提取为错误。
3. 检查数值一致性：严禁四舍五入数值；中文数量单位换算需先展开为纯数字再比对（万=10^4，亿=10^8）。
4. 检查编号连续性：层级编号是否跳号、重号。
5. 排除非数值文本：禁止修改嵌入在单词内部的类似数字字符，除非是明确编号。
6. 最小拆分原则：一句话内多个数值错误须拆分为多个 JSON 对象，"译文数值"/"替换锚点"精确到具体错误片段。

# 输出要求
仅输出 JSON 数组，不得包含说明文字。若无错误，输出 []。格式：
[
  {{
    "错误编号": "1",
    "原文上下文": "包含该数值的原文完整句",
    "译文上下文": "包含该数值的译文完整句",
    "原文数值": "原文提取的原文片段",
    "译文数值": "译文中提取的译文错误片段",
    "替换锚点": "译文中需要被替换的精确字符片段",
    "译文修改建议值": "修正后的译文片段（须来源于复制译文数值字段并修改其中的错误）",
    "is_source_consistent": "true或false — 译文是否忠实还原了原文数值（true=原文本身可能有问题，不需要改译文；false=翻译错误，需要改译文）",
    "错误类型": "数值错误",
    "修改理由": "简述违反的具体规则",
    "违反的规则": "规则条款"
  }}
]

# 输入数据
- 原文：{orig}
- 译文：{trans}
"""


def _compare_texts(original_text: str, translated_text: str) -> str:
    """调用 LLM 对比一段原文/译文，返回原始响应字符串。"""
    prompt = _PROMPT_TEMPLATE.format(orig=original_text, trans=translated_text)
    resp = client.chat.completions.create(
        model=_MODEL,
        messages=[
            {"role": "system", "content": "你是中译英译文合规审校员，只输出JSON数组，不做多余解释。"},
            {"role": "user", "content": prompt},
        ],
        temperature=0,
    )
    return resp.choices[0].message.content or ""


def _fix_suggestion_overlap(context: str, anchor: str, suggestion: str) -> str:
    """复用 clean_replace_duplicates.clean_suggestion 做去重裁剪。"""
    return clean_suggestion(context, anchor, suggestion)


def _compare_with_split(orig_txt: str, trans_txt: str, name: str) -> List[Dict]:
    """对一组原文/译文做分块对比，合并去重后返回扁平错误列表。"""
    if not orig_txt or not trans_txt:
        print(f"  ⚠️ {name}原文或译文为空，跳过")
        return []

    pairs = split_text_pair(orig_txt, trans_txt)
    num_parts = len(pairs)

    def _process_one(orig_chunk: str, trans_chunk: str) -> List[Dict]:
        raw = None
        for attempt in range(3):
            try:
                raw = _compare_texts(orig_chunk, trans_chunk)
                break
            except Exception as e:
                if attempt < 2:
                    wait = (attempt + 1) * 10
                    print(f"  ⚠️ 调用失败: {e}，{wait}秒后重试...")
                    time.sleep(wait)
                else:
                    print(f"  ❌ 调用 API 失败（已重试2次）: {e}")
        if not raw:
            return []
        parsed = parse_json_content(raw)
        if not isinstance(parsed, list):
            return []

        filtered = []
        for item in parsed:
            tran_val = (item.get("译文数值") or "").strip()
            tran_sug = (item.get("译文修改建议值") or "").strip()
            context_str = (item.get("译文上下文") or "").strip()
            anchor_str = (item.get("替换锚点") or tran_val).strip()
            fixed = _fix_suggestion_overlap(context_str, anchor_str, tran_sug)
            if fixed != tran_sug:
                item["译文修改建议值"] = fixed
                tran_sug = fixed
            if not tran_sug or tran_sug == tran_val:
                continue
            filtered.append(item)
        return filtered

    if num_parts <= 1:
        print(f"  📄 {name}文本较短（{_count_chars(orig_txt)} 字），直接对比")
        return _process_one(orig_txt, trans_txt)

    print(f"  📄 {name}文本较长（原文 {_count_chars(orig_txt)} 字），分割为 {num_parts} 块进行对比")

    all_errors: List[Dict] = []
    seen_keys = set()

    for i, (orig_chunk, trans_chunk, *_rest) in enumerate(pairs, 1):
        print(f"    --- {name} 第 {i}/{num_parts} 块 ---")
        if not orig_chunk.strip() or not trans_chunk.strip():
            print(f"      ⚠️ 块内容为空，跳过")
            continue
        items = _process_one(orig_chunk, trans_chunk)
        chunk_count = 0
        for item in items:
            dedup_key = (
                (item.get("译文数值") or "").strip(),
                (item.get("译文上下文") or "").strip()[:50],
            )
            if dedup_key in seen_keys:
                continue
            seen_keys.add(dedup_key)
            all_errors.append(item)
            chunk_count += 1
        print(f"      ✓ 本块发现 {chunk_count} 个问题，累计 {len(all_errors)} 个")

    for idx, item in enumerate(all_errors, 1):
        item["错误编号"] = str(idx)

    print(f"  📊 {name}合并完成: 共 {len(all_errors)} 个不重复问题")
    return all_errors


# =============================================================
# LLM 对比（单文件双语对照：一段文本内中英交替，无独立原文/译文入参）
# =============================================================

_BILINGUAL_PROMPT_TEMPLATE = """# 文件类型：双语对照或者多语种对照文件

# 角色
你是一名资深的文件翻译审校专家，精通输入文档中英文以及其他语种文件中的数值、金额、日期、编号层级的一致性核查。

# 任务
1. 严格对比输入文档中的原文与译文，识别并提取所有数值、单位、编号、日期翻译后译文与原文不一致的错误。
2. 对照原文检查译文出现错译、漏译、多译的情况，以【原文】为唯一准则。
3. 若译文符合规则且无错译，严禁将其放入 JSON 列表中输出。

# 审查规则（执行优先级最高）：
1. 原文主权原则：修改建议必须以【原文】为唯一准则，禁止引入原文中不存在的信息。
2. 提取纯粹性：严禁拼接、计算或修改数字。仅当译文数值与原文在含义、精确度或逻辑上不符时才提取为错误。
3. 检查数值一致性：严禁四舍五入数值；中文数量单位换算需先展开为纯数字再比对（万=10^4，亿=10^8）。
4. 检查编号连续性：层级编号是否跳号、重号或译文连续编号混用不同符号。
5. 排除非数值文本：禁止修改嵌入在单词内部的类似数字字符，除非是明确编号。
6. 最小拆分原则：一句话内多个数值错误须拆分为多个 JSON 对象，"译文数值"/"替换锚点"精确到具体错误片段。

# 输出要求
仅输出 JSON 数组，不得包含说明文字。若无错误，输出 []。格式：
[
  {{
    "错误编号": "1",
    "原文上下文": "包含该数值的原文完整句",
    "译文上下文": "包含该数值的译文完整句",
    "原文数值": "原文提取的原文片段",
    "译文数值": "译文中提取的译文错误片段",
    "替换锚点": "译文中需要被替换的精确字符片段",
    "译文修改建议值": "修正后的译文片段（须来源于复制译文数值字段并修改其中的错误）",
    "is_source_consistent": "true或false — 译文是否忠实还原了原文数值（true=原文本身可能有问题，不需要改译文；false=翻译错误，需要改译文）",
    "错误类型": "数值错误",
    "修改理由": "简述违反的具体规则",
    "违反的规则": "规则条款"
  }}
]

# 输入数据
- 输入文档：{doc}
"""


def _compare_bilingual_text(text: str) -> str:
    """调用 LLM 对比单份双语对照文本，返回原始响应字符串。"""
    prompt = _BILINGUAL_PROMPT_TEMPLATE.format(doc=text)
    resp = client.chat.completions.create(
        model=_MODEL,
        messages=[
            {"role": "system", "content": "你是中译英以及其他语种译文合规审校员，只输出JSON数组，不做多余解释。"},
            {"role": "user", "content": prompt},
        ],
        temperature=0,
    )
    return resp.choices[0].message.content or ""


def _compare_bilingual_with_split(text: str, name: str) -> List[Dict]:
    """对单文件双语对照文本按段落对分组分块，逐块直送 AI，合并去重。"""
    if not text or not text.strip():
        print(f"  ⚠️ {name}内容为空，跳过")
        return []

    chunks = split_bilingual_text(text)
    if not chunks:
        chunks = [text]
    num_parts = len(chunks)

    def _process_one(chunk: str) -> List[Dict]:
        raw = None
        for attempt in range(3):
            try:
                raw = _compare_bilingual_text(chunk)
                break
            except Exception as e:
                if attempt < 2:
                    wait = (attempt + 1) * 10
                    print(f"  ⚠️ 调用失败: {e}，{wait}秒后重试...")
                    time.sleep(wait)
                else:
                    print(f"  ❌ 调用 API 失败（已重试2次）: {e}")
        if not raw:
            return []
        parsed = parse_json_content(raw)
        if not isinstance(parsed, list):
            return []

        filtered = []
        for item in parsed:
            tran_val = (item.get("译文数值") or "").strip()
            tran_sug = (item.get("译文修改建议值") or "").strip()
            context_str = (item.get("译文上下文") or "").strip()
            anchor_str = (item.get("替换锚点") or tran_val).strip()
            fixed = _fix_suggestion_overlap(context_str, anchor_str, tran_sug)
            if fixed != tran_sug:
                item["译文修改建议值"] = fixed
                tran_sug = fixed
            if not tran_sug or tran_sug == tran_val:
                continue
            filtered.append(item)
        return filtered

    if num_parts <= 1:
        print(f"  📄 {name}文本较短（{_count_chars(text)} 字），直接对比")
        return _process_one(chunks[0] if chunks else text)

    print(f"  📄 {name}文本较长（{_count_chars(text)} 字），分割为 {num_parts} 块进行对比")

    all_errors: List[Dict] = []
    seen_keys = set()

    for i, chunk in enumerate(chunks, 1):
        print(f"    --- {name} 第 {i}/{num_parts} 块 ---")
        if not chunk.strip():
            print(f"      ⚠️ 块内容为空，跳过")
            continue
        items = _process_one(chunk)
        chunk_count = 0
        for item in items:
            dedup_key = (
                (item.get("译文数值") or "").strip(),
                (item.get("译文上下文") or "").strip()[:50],
            )
            if dedup_key in seen_keys:
                continue
            seen_keys.add(dedup_key)
            all_errors.append(item)
            chunk_count += 1
        print(f"      ✓ 本块发现 {chunk_count} 个问题，累计 {len(all_errors)} 个")

    for idx, item in enumerate(all_errors, 1):
        item["错误编号"] = str(idx)

    print(f"  📊 {name}合并完成: 共 {len(all_errors)} 个不重复问题")
    return all_errors


# =============================================================
# 主流程：提取 + AI 对比 → 扁平错误列表（按区域分组）
# =============================================================

def run_legacy_comparison(src_docx_path: str, tgt_docx_path: str = None,
                          bilingual_mode: bool = False) -> Dict[str, List[Dict]]:
    """
    旧版流程：整篇提取 → 分块直送 AI → 合并去重。

    两种模式：
      双文件模式（bilingual_mode=False，默认）：
        src_docx_path/tgt_docx_path 为两个独立文件，
        按 split_text_pair 的比例对齐切块，双段式 prompt（原文/译文分开传入）。
      单文件双语对照模式（bilingual_mode=True）：
        只需 src_docx_path 一个文件（一篇文档内中英段落交替排列），
        按 split_bilingual_text 的中英分组切块，单段式 prompt（整块文本一起传入）。

    返回 {"body": [...], "header": [...], "footer": [...]}，
    每个 error dict 额外带上 "_region" 字段（body/header/footer），供写回定位。
    """
    if bilingual_mode:
        print("📖 提取双语对照文档...")
        body, header, footer = extract_channels(src_docx_path)

        channels = [
            ("body", "正文", body),
            ("header", "页眉", header),
            ("footer", "页脚", footer),
        ]

        result: Dict[str, List[Dict]] = {}
        for region, label, text in channels:
            print(f"\n{'='*60}\n📄 {label}检查（旧版·单文件双语对照模式）\n{'='*60}")
            errors = _compare_bilingual_with_split(text, label)
            for e in errors:
                e["_region"] = region
            result[region] = errors
        return result

    if not tgt_docx_path:
        raise ValueError("双文件模式（bilingual_mode=False）必须提供 tgt_docx_path")

    print("📖 提取原文...")
    src_body, src_header, src_footer = extract_channels(src_docx_path)
    print("📖 提取译文...")
    tgt_body, tgt_header, tgt_footer = extract_channels(tgt_docx_path)

    channels = [
        ("body", "正文", src_body, tgt_body),
        ("header", "页眉", src_header, tgt_header),
        ("footer", "页脚", src_footer, tgt_footer),
    ]

    result: Dict[str, List[Dict]] = {}
    for region, label, orig_txt, tran_txt in channels:
        print(f"\n{'='*60}\n📄 {label}检查（旧版·双文件模式）\n{'='*60}")
        errors = _compare_with_split(orig_txt, tran_txt, label)
        for e in errors:
            e["_region"] = region
        result[region] = errors

    return result


# =============================================================
# 写回：仅用锚点/上下文匹配（降级但贴合旧流程，不依赖 para_index）
# =============================================================

def apply_legacy_errors(errors_by_region: Dict[str, List[Dict]],
                        docx: object,
                        revision_manager: object,
                        doc_path: str = None) -> Tuple[list, int, int]:
    """
    将 run_legacy_comparison() 的扁平错误列表写入 Track Changes。

    与 main.py::apply_revisions_from_ai_map 的区别：
    - 不使用 para_index / prev_tgt / next_tgt 夹逼定位（旧流程没有逐段对齐信息）
    - 仅依赖 替换锚点 / 译文上下文 做定位（replace_and_revise_in_docx 的策略1/2/3-6）
    """
    from replace_revision import replace_and_revise_in_docx

    tasks = []
    skipped = []

    for region, errors in errors_by_region.items():
        for err_no, err in enumerate(errors, 1):
            isc = err.get("is_source_consistent")
            if isc is True or (isinstance(isc, str) and isc.lower() == "true"):
                skipped.append((region, err_no, "原译一致，疑似原文问题")); continue
            old_val = (err.get("替换锚点") or err.get("译文数值") or "").strip()
            new_val = (err.get("译文修改建议值") or "").strip()
            context = (err.get("译文上下文") or "").strip()
            anchor = (err.get("替换锚点") or "").strip()
            reason = err.get("修改理由", "") or err.get("错误类型", "")

            if not old_val or not new_val or old_val == new_val:
                skipped.append((region, err_no, "字段缺失或无变化")); continue

            tasks.append((region, err_no, old_val, new_val, context, anchor, reason))

    if skipped:
        print(f"  [过滤] {len(skipped)} 条（字段缺失或无变化）")

    # 长文本优先，避免短锚点抢先替换
    tasks.sort(key=lambda t: len(t[2]), reverse=True)

    success, failed = 0, 0
    for i, (region, err_no, old_val, new_val, context, anchor, reason) in enumerate(tasks, 1):
        ok, strategy = replace_and_revise_in_docx(
            doc=docx, old_value=old_val, new_value=new_val, reason=reason,
            revision_manager=revision_manager, context=context, anchor_text=anchor,
            region=region, doc_path=doc_path,
        )
        if ok:
            success += 1
            print(f"  ✅ [{i:>2}/{len(tasks)}] [{region}] '{old_val}' → '{new_val}'  ({strategy})")
        else:
            failed += 1
            print(f"  ⚠️  [{i:>2}/{len(tasks)}] [{region}] 未找到: '{old_val}'  ({strategy})")

    return tasks, success, failed


# =============================================================
# 报告：扁平错误列表 → Excel（每区域一个 sheet）
# =============================================================

def generate_legacy_report(errors_by_region: Dict[str, List[Dict]],
                           output_path: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    _cols = ["错误编号", "错误类型", "原文数值", "译文数值", "译文修改建议值",
             "修改理由", "违反的规则", "原文上下文", "译文上下文",
             "替换锚点", "is_source_consistent"]
    _widths = [8, 15, 20, 20, 20, 30, 30, 40, 40, 20, 12]
    _region_label = {"body": "正文", "header": "页眉", "footer": "页脚"}

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    for region, errors in errors_by_region.items():
        if not errors:
            continue
        ws = wb.create_sheet(_region_label.get(region, region))
        ws.append(_cols)
        fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for i, w in enumerate(_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for err in errors:
            ws.append([err.get(c, "") for c in _cols])

    if not wb.sheetnames:
        ws = wb.create_sheet("无错误")
        ws.append(["共发现 0 个需要修改的对象"])

    wb.save(output_path)
    print(f"✅ 旧版流程报告: {output_path}")
    return output_path
